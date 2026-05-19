import streamlit as st
import pandas as pd

from data.loader import PainelData
from data.editor import (
    save_config, save_fixos, save_pj_parcelas,
    save_parcelas_pessoais, save_tomadores, save_faturamento,
    save_pontuais_mes, CONFIG_LABELS,
)
from config import MONTHS, MONTH_LABELS


def _save_btn(label: str = "💾 Salvar"):
    return st.button(label, type="primary", use_container_width=True)


def _flush(xlsx_path: str):
    st.cache_data.clear()
    st.success("Salvo! Recarregando dados…")
    st.rerun()


def render(data: PainelData):
    st.header("✏️ Editar Dados")
    st.caption("Todas as alterações são gravadas diretamente no arquivo .xlsx. O original é modificado.")

    xlsx_path = st.session_state.get("xlsx_path_current", "")
    if not xlsx_path:
        st.error("Caminho do xlsx não encontrado. Verifique o sidebar.")
        return

    tabs = st.tabs([
        "⚙️ Configurações",
        "🔒 Fixos Pessoais",
        "📦 Parc. PJ",
        "💳 Parc. Pessoais",
        "🏢 Tomadores",
        "💼 Faturamento PJ",
        "📅 Pontuais do Mês",
    ])

    # ── Configurações ─────────────────────────────────────────────────────────
    with tabs[0]:
        st.subheader("⚙️ Configurações")
        cfg = data.config or {}

        col1, col2 = st.columns(2)
        inputs = {}

        with col1:
            st.markdown("**Saldos e Provisão**")
            inputs["saldo_atual"] = st.number_input(
                "Saldo atual em conta (R$)", value=float(cfg.get("saldo_atual", 0) or 0), step=100.0, format="%.2f"
            )
            inputs["provisao_fiscal"] = st.number_input(
                "Provisão fiscal acumulada (R$)", value=float(cfg.get("provisao_fiscal", 0) or 0), step=100.0, format="%.2f"
            )
            inputs["provisao_pct"] = st.number_input(
                "% Provisão fiscal", value=float(cfg.get("provisao_pct", 0) or 0), step=0.01, format="%.4f",
                help="Ex: 0.25 para 25%"
            )

            st.markdown("**Alertas**")
            inputs["limite_alerta"] = st.number_input(
                "Limite de alerta (R$)", value=float(cfg.get("limite_alerta", 0) or 0), step=100.0, format="%.2f"
            )
            inputs["reserva_meses"] = st.number_input(
                "Reserva de emergência alvo (meses)", value=float(cfg.get("reserva_meses", 3) or 3), step=0.5, format="%.1f"
            )

        with col2:
            st.markdown("**Estimativas Mensais**")
            inputs["sm24_estimativa"] = st.number_input(
                "SM24 (R$)", value=float(cfg.get("sm24_estimativa", 0) or 0), step=100.0, format="%.2f"
            )
            inputs["amor_saude"] = st.number_input(
                "Amor Saúde (R$)", value=float(cfg.get("amor_saude", 0) or 0), step=100.0, format="%.2f"
            )
            inputs["pucpr"] = st.number_input(
                "PUCPR (R$)", value=float(cfg.get("pucpr", 0) or 0), step=100.0, format="%.2f"
            )
            inputs["naturalles"] = st.number_input(
                "Naturalles (R$)", value=float(cfg.get("naturalles", 0) or 0), step=100.0, format="%.2f"
            )
            inputs["bolsa_residencia"] = st.number_input(
                "Bolsa Residência (R$)", value=float(cfg.get("bolsa_residencia", 0) or 0), step=100.0, format="%.2f"
            )

            st.markdown("**Alíquotas de Impostos**")
            inputs["pis"] = st.number_input("PIS", value=float(cfg.get("pis", 0) or 0), step=0.0001, format="%.4f")
            inputs["cofins"] = st.number_input("COFINS", value=float(cfg.get("cofins", 0) or 0), step=0.0001, format="%.4f")
            inputs["irpj"] = st.number_input("IRPJ", value=float(cfg.get("irpj", 0) or 0), step=0.0001, format="%.4f")
            inputs["csll"] = st.number_input("CSLL", value=float(cfg.get("csll", 0) or 0), step=0.0001, format="%.4f")
            inputs["iss"] = st.number_input("ISS Maringá", value=float(cfg.get("iss", 0) or 0), step=0.0001, format="%.4f")

        if _save_btn():
            try:
                save_config(xlsx_path, inputs)
                _flush(xlsx_path)
            except Exception as e:
                st.error(f"Erro ao salvar: {e}")

    # ── Fixos Pessoais ────────────────────────────────────────────────────────
    with tabs[1]:
        st.subheader("🔒 Gastos Fixos Pessoais")
        st.caption("Máximo 13 itens (linhas 5-17). Colunas de fórmula são preservadas.")

        fixos_df = data.fixos.copy() if not data.fixos.empty else pd.DataFrame(
            columns=["Descrição", "Categoria", "Valor", "Status", "Obs"]
        )

        for col in ["Descrição", "Categoria", "Obs", "Status"]:
            if col not in fixos_df.columns:
                fixos_df[col] = ""
        if "Valor" not in fixos_df.columns:
            fixos_df["Valor"] = 0.0

        # Pad to 13 rows
        while len(fixos_df) < 13:
            fixos_df = pd.concat([fixos_df, pd.DataFrame([{c: "" for c in fixos_df.columns}])], ignore_index=True)

        edited = st.data_editor(
            fixos_df[["Descrição", "Categoria", "Valor", "Status", "Obs"]].head(13),
            use_container_width=True,
            hide_index=False,
            num_rows="fixed",
            column_config={
                "Valor": st.column_config.NumberColumn("Valor (R$)", format="R$ %.2f"),
                "Status": st.column_config.SelectboxColumn("Status", options=["Ativo", "Pausado", "Cancelado", ""]),
            },
            key="editor_fixos",
        )

        if _save_btn():
            try:
                save_fixos(xlsx_path, edited)
                _flush(xlsx_path)
            except Exception as e:
                st.error(f"Erro ao salvar: {e}")

    # ── PJ Parcelamentos ──────────────────────────────────────────────────────
    with tabs[2]:
        st.subheader("📦 Parcelamentos PJ")
        st.caption("Máximo 13 itens. Colunas 'Mês Final' e 'Total' são calculadas automaticamente pela planilha.")

        pj_df = data.pj_parcelas_config.copy() if not data.pj_parcelas_config.empty else pd.DataFrame(
            columns=["Descrição", "Parcela", "N Parcelas", "Mês Início", "Ano Início", "Status"]
        )

        for col in ["Descrição", "Status"]:
            if col not in pj_df.columns:
                pj_df[col] = ""
        for col in ["Parcela", "N Parcelas", "Mês Início", "Ano Início"]:
            if col not in pj_df.columns:
                pj_df[col] = 0

        while len(pj_df) < 13:
            pj_df = pd.concat([pj_df, pd.DataFrame([{c: "" for c in pj_df.columns}])], ignore_index=True)

        edited = st.data_editor(
            pj_df[["Descrição", "Parcela", "N Parcelas", "Mês Início", "Ano Início", "Status"]].head(13),
            use_container_width=True,
            hide_index=False,
            num_rows="fixed",
            column_config={
                "Parcela": st.column_config.NumberColumn("Parcela (R$)", format="R$ %.2f"),
                "N Parcelas": st.column_config.NumberColumn("Nº Parcelas", format="%d"),
                "Mês Início": st.column_config.NumberColumn("Mês Início", format="%d", min_value=1, max_value=12),
                "Ano Início": st.column_config.NumberColumn("Ano Início", format="%d"),
                "Status": st.column_config.SelectboxColumn("Status", options=["Ativo", "Quitado", "Suspenso", ""]),
            },
            key="editor_pj_parcelas",
        )

        if _save_btn():
            try:
                save_pj_parcelas(xlsx_path, edited)
                _flush(xlsx_path)
            except Exception as e:
                st.error(f"Erro ao salvar: {e}")

    # ── Parcelas Pessoais ─────────────────────────────────────────────────────
    with tabs[3]:
        st.subheader("💳 Parcelas Pessoais")
        st.caption("Máximo 20 itens. Colunas 'Mês Final' e 'Total' são calculadas automaticamente pela planilha.")

        parc_df = data.parcelas_pessoais_config.copy() if not data.parcelas_pessoais_config.empty else pd.DataFrame(
            columns=["Descrição", "Categoria", "Parcela", "N Parcelas", "Mês Início", "Ano Início", "Obs"]
        )

        for col in ["Descrição", "Categoria", "Obs"]:
            if col not in parc_df.columns:
                parc_df[col] = ""
        for col in ["Parcela", "N Parcelas", "Mês Início", "Ano Início"]:
            if col not in parc_df.columns:
                parc_df[col] = 0

        while len(parc_df) < 20:
            parc_df = pd.concat([parc_df, pd.DataFrame([{c: "" for c in parc_df.columns}])], ignore_index=True)

        edited = st.data_editor(
            parc_df[["Descrição", "Categoria", "Parcela", "N Parcelas", "Mês Início", "Ano Início", "Obs"]].head(20),
            use_container_width=True,
            hide_index=False,
            num_rows="fixed",
            column_config={
                "Parcela": st.column_config.NumberColumn("Parcela (R$)", format="R$ %.2f"),
                "N Parcelas": st.column_config.NumberColumn("Nº Parcelas", format="%d"),
                "Mês Início": st.column_config.NumberColumn("Mês Início", format="%d", min_value=1, max_value=12),
                "Ano Início": st.column_config.NumberColumn("Ano Início", format="%d"),
            },
            key="editor_parcelas_pessoais",
        )

        if _save_btn():
            try:
                save_parcelas_pessoais(xlsx_path, edited)
                _flush(xlsx_path)
            except Exception as e:
                st.error(f"Erro ao salvar: {e}")

    # ── Tomadores ─────────────────────────────────────────────────────────────
    with tabs[4]:
        st.subheader("🏢 Tomadores de Serviço")
        st.caption("Máximo 6 tomadores (linhas 5-10).")

        # Load directly from xlsx
        try:
            from openpyxl import load_workbook as _lwb
            _wb = _lwb(xlsx_path, data_only=True)
            _ws = _wb["Tomadores"]
            _rows = []
            for r in range(5, 11):
                _rows.append({
                    "Nome": _ws.cell(r, 1).value or "",
                    "Tipo": _ws.cell(r, 2).value or "",
                    "Retém": _ws.cell(r, 3).value or "",
                    "Tipo receita": _ws.cell(r, 4).value or "",
                    "Obs": _ws.cell(r, 5).value or "",
                })
            tom_df = pd.DataFrame(_rows)
        except Exception:
            tom_df = pd.DataFrame(columns=["Nome", "Tipo", "Retém", "Tipo receita", "Obs"])

        edited = st.data_editor(
            tom_df,
            use_container_width=True,
            hide_index=False,
            num_rows="fixed",
            column_config={
                "Tipo": st.column_config.SelectboxColumn("Tipo", options=["PF", "PJ", ""]),
                "Retém": st.column_config.SelectboxColumn("Retém ISS?", options=["Sim", "Não", ""]),
                "Tipo receita": st.column_config.SelectboxColumn(
                    "Tipo Receita", options=["Plantão", "Consultório", "Salário", "Outros", ""]
                ),
            },
            key="editor_tomadores",
        )

        if _save_btn():
            try:
                save_tomadores(xlsx_path, edited)
                _flush(xlsx_path)
            except Exception as e:
                st.error(f"Erro ao salvar: {e}")

    # ── Faturamento PJ ────────────────────────────────────────────────────────
    with tabs[5]:
        st.subheader("💼 Faturamento PJ por Cliente")
        st.caption("Linhas = meses (Mai-Dez), Colunas = até 6 clientes. Edite os valores faturados.")

        fat_df = data.faturamento.copy() if not data.faturamento.empty else pd.DataFrame()

        if fat_df.empty:
            st.info("Nenhum dado de faturamento encontrado no xlsx.")
        else:
            month_col = "Mês" if "Mês" in fat_df.columns else fat_df.columns[0]
            client_cols = [c for c in fat_df.columns if c != month_col]

            col_cfg = {
                c: st.column_config.NumberColumn(c, format="R$ %.2f") for c in client_cols
            }
            col_cfg[month_col] = st.column_config.TextColumn(month_col, disabled=True)

            edited = st.data_editor(
                fat_df,
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                column_config=col_cfg,
                key="editor_faturamento",
            )

            if _save_btn():
                try:
                    save_faturamento(xlsx_path, edited)
                    _flush(xlsx_path)
                except Exception as e:
                    st.error(f"Erro ao salvar: {e}")

    # ── Pontuais do Mês ───────────────────────────────────────────────────────
    with tabs[6]:
        st.subheader("📅 Gastos Pontuais por Mês")

        sel_idx = st.selectbox(
            "Mês:",
            options=range(len(MONTHS)),
            format_func=lambda i: MONTH_LABELS[i],
            key="editor_month_sel",
        )
        month_key = MONTHS[sel_idx]
        month_data = data.months.get(month_key)

        pont_df = month_data.pontuais.copy() if (month_data and not month_data.pontuais.empty) else pd.DataFrame(
            columns=["Descrição", "Categoria", "Valor", "Status"]
        )

        for col in ["Descrição", "Categoria", "Status"]:
            if col not in pont_df.columns:
                pont_df[col] = ""
        if "Valor" not in pont_df.columns:
            pont_df["Valor"] = 0.0

        from data.reconciler import ALL_CATEGORIES

        edited = st.data_editor(
            pont_df[["Descrição", "Categoria", "Valor", "Status"]],
            use_container_width=True,
            hide_index=False,
            num_rows="dynamic",
            column_config={
                "Valor": st.column_config.NumberColumn("Valor (R$)", format="R$ %.2f"),
                "Categoria": st.column_config.SelectboxColumn("Categoria", options=ALL_CATEGORIES),
                "Status": st.column_config.SelectboxColumn(
                    "Status", options=["Confirmado", "Previsto", "Cancelado", ""]
                ),
            },
            key=f"editor_pontuais_{month_key}",
        )

        st.caption(f"Máximo 16 itens por mês. Atualmente: {len(edited)} linha(s).")

        if _save_btn():
            try:
                save_pontuais_mes(xlsx_path, month_key, edited)
                _flush(xlsx_path)
            except Exception as e:
                st.error(f"Erro ao salvar: {e}")
