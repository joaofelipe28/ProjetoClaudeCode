"""
Load and save financial goals (Metas) from/to the xlsx.
Creates the "Metas" sheet automatically if it doesn't exist.
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

METAS_SHEET = "Metas"
METAS_HEADERS = ["Descrição", "Categoria", "Valor Alvo", "Valor Atual", "Prazo", "Status"]
METAS_COL_MAP = {h: i + 1 for i, h in enumerate(METAS_HEADERS)}
METAS_DATA_START = 2


def _ensure_metas_sheet(wb):
    if METAS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(METAS_SHEET)
        header_fill = PatternFill("solid", start_color="1C2333")
        for i, h in enumerate(METAS_HEADERS, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True, color="FAFAFA")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 16
    return wb


def load_metas(xlsx_path: str) -> pd.DataFrame:
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        if METAS_SHEET not in wb.sheetnames:
            return pd.DataFrame(columns=METAS_HEADERS)
        ws = wb[METAS_SHEET]
        rows = []
        for r in range(METAS_DATA_START, ws.max_row + 1):
            desc = ws.cell(r, 1).value
            if not desc:
                continue
            rows.append({h: ws.cell(r, METAS_COL_MAP[h]).value for h in METAS_HEADERS})
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=METAS_HEADERS)
    except Exception:
        return pd.DataFrame(columns=METAS_HEADERS)


def save_metas(xlsx_path: str, df: pd.DataFrame):
    wb = load_workbook(xlsx_path, data_only=False)
    wb = _ensure_metas_sheet(wb)
    ws = wb[METAS_SHEET]

    # Clear existing data rows
    for r in range(METAS_DATA_START, ws.max_row + 1):
        for c in range(1, len(METAS_HEADERS) + 1):
            ws.cell(row=r, column=c).value = None

    # Write new data
    for i, (_, row) in enumerate(df.iterrows()):
        r = METAS_DATA_START + i
        for h, col_num in METAS_COL_MAP.items():
            val = row.get(h, None)
            if val is not None and not (isinstance(val, float) and pd.isna(val)):
                ws.cell(row=r, column=col_num).value = val

    wb.save(xlsx_path)
