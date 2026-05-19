"""
Load and save investment positions from/to the xlsx.
Creates the "Investimentos" sheet automatically if it doesn't exist.
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

INV_SHEET = "Investimentos"
INV_HEADERS = ["Nome", "Tipo", "Instituição", "Valor Aplicado", "Saldo Atual", "Taxa a.a. (%)", "Vencimento", "Liquidez", "Obs"]
INV_COL_MAP = {h: i + 1 for i, h in enumerate(INV_HEADERS)}
INV_DATA_START = 2


def _ensure_inv_sheet(wb):
    if INV_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(INV_SHEET)
        header_fill = PatternFill("solid", start_color="1C2333")
        for i, h in enumerate(INV_HEADERS, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True, color="FAFAFA")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        widths = [28, 18, 18, 16, 16, 14, 14, 14, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    return wb


def load_investments(xlsx_path: str) -> pd.DataFrame:
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        if INV_SHEET not in wb.sheetnames:
            return pd.DataFrame(columns=INV_HEADERS)
        ws = wb[INV_SHEET]
        rows = []
        for r in range(INV_DATA_START, ws.max_row + 1):
            nome = ws.cell(r, 1).value
            if not nome:
                continue
            rows.append({h: ws.cell(r, INV_COL_MAP[h]).value for h in INV_HEADERS})
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=INV_HEADERS)
    except Exception:
        return pd.DataFrame(columns=INV_HEADERS)


def save_investments(xlsx_path: str, df: pd.DataFrame):
    wb = load_workbook(xlsx_path, data_only=False)
    wb = _ensure_inv_sheet(wb)
    ws = wb[INV_SHEET]

    for r in range(INV_DATA_START, ws.max_row + 1):
        for c in range(1, len(INV_HEADERS) + 1):
            ws.cell(row=r, column=c).value = None

    for i, (_, row) in enumerate(df.iterrows()):
        r = INV_DATA_START + i
        for h, col_num in INV_COL_MAP.items():
            val = row.get(h, None)
            if val is not None and not (isinstance(val, float) and pd.isna(val)):
                ws.cell(row=r, column=col_num).value = val

    wb.save(xlsx_path)


# ── Investment monthly history ────────────────────────────────────────────────

INV_HIST_SHEET = "Inv_Historico"
INV_HIST_HEADERS = ["Mês", "Aporte (R$)", "Saldo Final (R$)", "Rentabilidade (%)", "Obs"]
INV_HIST_COL_MAP = {h: i + 1 for i, h in enumerate(INV_HIST_HEADERS)}
INV_HIST_DATA_START = 2


def _ensure_inv_hist_sheet(wb):
    if INV_HIST_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(INV_HIST_SHEET)
        header_fill = PatternFill("solid", start_color="1C2333")
        for i, h in enumerate(INV_HIST_HEADERS, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True, color="FAFAFA")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 24
    return wb


def load_inv_historico(xlsx_path: str) -> pd.DataFrame:
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        if INV_HIST_SHEET not in wb.sheetnames:
            return pd.DataFrame(columns=INV_HIST_HEADERS)
        ws = wb[INV_HIST_SHEET]
        rows = []
        for r in range(INV_HIST_DATA_START, ws.max_row + 1):
            mes = ws.cell(r, 1).value
            if not mes:
                continue
            rows.append({h: ws.cell(r, INV_HIST_COL_MAP[h]).value for h in INV_HIST_HEADERS})
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=INV_HIST_HEADERS)
    except Exception:
        return pd.DataFrame(columns=INV_HIST_HEADERS)


def save_inv_historico(xlsx_path: str, df: pd.DataFrame):
    wb = load_workbook(xlsx_path, data_only=False)
    wb = _ensure_inv_hist_sheet(wb)
    ws = wb[INV_HIST_SHEET]

    current_max = max(ws.max_row, INV_HIST_DATA_START + len(df) + 1)
    for r in range(INV_HIST_DATA_START, current_max + 1):
        for c in range(1, len(INV_HIST_HEADERS) + 1):
            ws.cell(row=r, column=c).value = None

    for i, (_, row) in enumerate(df.iterrows()):
        r = INV_HIST_DATA_START + i
        for h, col_num in INV_HIST_COL_MAP.items():
            val = row.get(h, None)
            if val is not None and not (isinstance(val, float) and pd.isna(val)):
                ws.cell(row=r, column=col_num).value = val

    wb.save(xlsx_path)
