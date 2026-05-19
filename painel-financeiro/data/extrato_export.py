"""
Write reconciled data back to the xlsx file.
Never overwrites the original — always returns bytes for download.
"""
import io
from typing import List, Dict
from openpyxl import load_workbook

from config import MONTHS


def _find_row_by_desc(ws, target_desc: str, desc_col: int = 1, search_rows: int = 120) -> int:
    """Find the row number in a worksheet where col desc_col contains target_desc (fuzzy)."""
    import difflib
    best_row = -1
    best_score = 0.0
    for r in range(1, search_rows):
        cell_val = ws.cell(row=r, column=desc_col).value
        if not cell_val:
            continue
        score = difflib.SequenceMatcher(None, str(cell_val).lower(), target_desc.lower()).ratio()
        if score > best_score:
            best_score = score
            best_row = r
    return best_row if best_score >= 0.5 else -1


def _find_section_end(ws, section_row: int, max_search: int = 40) -> int:
    """Find the last empty or total row after a section header."""
    for r in range(section_row + 2, section_row + max_search):
        v = str(ws.cell(row=r, column=1).value or "").strip()
        if not v:
            return r
        if "TOTAL" in v.upper() or "SALDO" in v.upper():
            return r
    return section_row + max_search


def _find_pontuais_section(ws, max_row: int = 150) -> int:
    """Find the row where section 7️⃣ (Gastos Pontuais) starts."""
    for r in range(1, max_row):
        v = str(ws.cell(row=r, column=1).value or "")
        if "7️⃣" in v:
            return r
    return -1


def export_xlsx(
    original_path: str,
    month_key: str,
    revenues_updates: List[Dict],   # [{"descricao": ..., "realizado": ..., "status": "Pago"}]
    pontuais_to_add: List[Dict],    # [{"data": ..., "descricao": ..., "categoria": ..., "valor": ..., "status": "Confirmado"}]
) -> bytes:
    """
    Load original xlsx, apply updates, return updated file as bytes.

    revenues_updates: update Realizado + Status columns in the receitas section.
    pontuais_to_add: append rows to section 7️⃣ in the month sheet.
    """
    wb = load_workbook(original_path, data_only=False)

    if month_key not in wb.sheetnames:
        wb_bytes = io.BytesIO()
        wb.save(wb_bytes)
        return wb_bytes.getvalue()

    ws = wb[month_key]

    # ── Update Realizado + Status for revenues ──────────────────────────────
    for update in revenues_updates:
        desc = update.get("descricao", "")
        realizado = update.get("realizado")
        status = update.get("status", "Pago")

        row = _find_row_by_desc(ws, desc, desc_col=1)
        if row == -1:
            continue

        # Receitas section: Descrição=col1, Tipo=2, Data=3, Previsto=4, Realizado=5, Status=6
        if realizado is not None:
            ws.cell(row=row, column=5).value = realizado
        ws.cell(row=row, column=6).value = status

    # ── Append pontuais rows to section 7️⃣ ────────────────────────────────
    if pontuais_to_add:
        sec_row = _find_pontuais_section(ws)
        if sec_row != -1:
            insert_row = _find_section_end(ws, sec_row)

            # Shift rows down to make room
            ws.insert_rows(insert_row, amount=len(pontuais_to_add))

            for i, p in enumerate(pontuais_to_add):
                r = insert_row + i
                ws.cell(row=r, column=1).value = p.get("descricao", "")
                ws.cell(row=r, column=2).value = p.get("categoria", "")
                ws.cell(row=r, column=3).value = p.get("valor", 0)
                ws.cell(row=r, column=4).value = p.get("status", "Confirmado")
                ws.cell(row=r, column=5).value = p.get("data", "")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
