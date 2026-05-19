"""
Write edited data back to the xlsx file.
Each function writes only to INPUT cells, never touching formulas.
"""
import io
from typing import Optional
import pandas as pd
from openpyxl import load_workbook


def _open_wb(xlsx_path: str):
    return load_workbook(xlsx_path, data_only=False)


def _save_wb(wb, xlsx_path: str):
    wb.save(xlsx_path)


def _clear_range(ws, row_start: int, row_end: int, cols: list):
    """Set cells to None in the given row range and column list."""
    for r in range(row_start, row_end + 1):
        for c in cols:
            ws.cell(row=r, column=c).value = None


# ── Configurações ────────────────────────────────────────────────────────────

CONFIG_LABELS = {
    "saldo_atual":      ("Saldo atual em conta", 5),
    "provisao_fiscal":  ("Provisão fiscal acumulada", 6),
    "sm24_estimativa":  ("Estimativa mensal de plantão SM24", 9),
    "amor_saude":       ("Estimativa mensal Amor Saúde", 10),
    "pucpr":            ("Estimativa mensal PUCPR", 11),
    "naturalles":       ("Estimativa mensal Naturalles", 12),
    "bolsa_residencia": ("Bolsa residência", 13),
    "pis":              ("PIS", 16),
    "cofins":           ("COFINS", 17),
    "irpj":             ("IRPJ", 18),
    "csll":             ("CSLL", 19),
    "iss":              ("ISS Maringá", 20),
    "provisao_pct":     ("% provisão fiscal", 21),
    "limite_alerta":    ("Limite de alerta", 24),
    "reserva_meses":    ("Reserva de emergência alvo", 25),
}


def save_config(xlsx_path: str, values: dict):
    """
    values: dict of key -> new value (same keys as CONFIG_LABELS).
    Finds each row by label substring match and updates col 2.
    """
    wb = _open_wb(xlsx_path)
    ws = wb["Configurações"]

    for key, new_val in values.items():
        if key not in CONFIG_LABELS:
            continue
        _, expected_row = CONFIG_LABELS[key]
        ws.cell(row=expected_row, column=2).value = new_val

    _save_wb(wb, xlsx_path)


# ── Fixos Pessoais ────────────────────────────────────────────────────────────
# Data rows: 5-17, Input cols: 1(Descrição), 2(Categoria), 3(Valor), 4(Status), 5(Obs)

FIXOS_DATA_START = 5
FIXOS_DATA_END = 17
FIXOS_INPUT_COLS = [1, 2, 3, 4, 5]


def save_fixos(xlsx_path: str, df: pd.DataFrame):
    """
    df columns: Descrição, Categoria, Valor, Status, Obs
    Clears rows 5-17 and rewrites with df data.
    """
    wb = _open_wb(xlsx_path)
    ws = wb["Fixos Pessoais"]
    _clear_range(ws, FIXOS_DATA_START, FIXOS_DATA_END, FIXOS_INPUT_COLS)

    col_map = {"Descrição": 1, "Categoria": 2, "Valor": 3, "Status": 4, "Obs": 5}
    for i, (_, row) in enumerate(df.iterrows()):
        if i >= (FIXOS_DATA_END - FIXOS_DATA_START + 1):
            break
        r = FIXOS_DATA_START + i
        for col_name, col_num in col_map.items():
            val = row.get(col_name, None)
            ws.cell(row=r, column=col_num).value = None if (pd.isna(val) if val is not None else False) else val

    _save_wb(wb, xlsx_path)


# ── PJ Parcelamentos ──────────────────────────────────────────────────────────
# Data rows: 5-17, Input cols: 1(Desc), 2(Valor parcela), 3(Nº parcelas), 4(Mês ini), 5(Ano ini), 8(Status)
# Cols 6(Mês final) and 7(Total) are FORMULAS — do NOT touch

PJ_PARC_DATA_START = 5
PJ_PARC_DATA_END = 17
PJ_PARC_INPUT_COLS = [1, 2, 3, 4, 5, 8]


def save_pj_parcelas(xlsx_path: str, df: pd.DataFrame):
    """
    df columns: Descrição, Parcela, N Parcelas, Mês Início, Ano Início, Status, Obs
    """
    wb = _open_wb(xlsx_path)
    ws = wb["PJ Parcelamentos"]
    _clear_range(ws, PJ_PARC_DATA_START, PJ_PARC_DATA_END, PJ_PARC_INPUT_COLS)

    col_map = {
        "Descrição": 1, "Parcela": 2, "N Parcelas": 3,
        "Mês Início": 4, "Ano Início": 5, "Status": 8,
    }
    for i, (_, row) in enumerate(df.iterrows()):
        if i >= (PJ_PARC_DATA_END - PJ_PARC_DATA_START + 1):
            break
        r = PJ_PARC_DATA_START + i
        for col_name, col_num in col_map.items():
            val = row.get(col_name, None)
            ws.cell(row=r, column=col_num).value = None if (pd.isna(val) if val is not None else False) else val

    _save_wb(wb, xlsx_path)


# ── Parcelas Pessoais ─────────────────────────────────────────────────────────
# Data rows: 5-24, Input cols: 1(Desc), 2(Cat), 3(Valor), 4(N parcelas), 5(Mês ini), 6(Ano ini), 9(Obs)
# Cols 7(Mês final) and 8(Total) are FORMULAS

PARC_PES_DATA_START = 5
PARC_PES_DATA_END = 24
PARC_PES_INPUT_COLS = [1, 2, 3, 4, 5, 6, 9]


def save_parcelas_pessoais(xlsx_path: str, df: pd.DataFrame):
    """
    df columns: Descrição, Categoria, Parcela, N Parcelas, Mês Início, Ano Início, Obs
    """
    wb = _open_wb(xlsx_path)
    ws = wb["Parcelas Pessoais"]
    _clear_range(ws, PARC_PES_DATA_START, PARC_PES_DATA_END, PARC_PES_INPUT_COLS)

    col_map = {
        "Descrição": 1, "Categoria": 2, "Parcela": 3,
        "N Parcelas": 4, "Mês Início": 5, "Ano Início": 6, "Obs": 9,
    }
    for i, (_, row) in enumerate(df.iterrows()):
        if i >= (PARC_PES_DATA_END - PARC_PES_DATA_START + 1):
            break
        r = PARC_PES_DATA_START + i
        for col_name, col_num in col_map.items():
            val = row.get(col_name, None)
            ws.cell(row=r, column=col_num).value = None if (pd.isna(val) if val is not None else False) else val

    _save_wb(wb, xlsx_path)


# ── Tomadores ─────────────────────────────────────────────────────────────────
# Data rows: 5-10, Input cols: 1-5

TOMADORES_DATA_START = 5
TOMADORES_DATA_END = 10
TOMADORES_INPUT_COLS = [1, 2, 3, 4, 5]


def save_tomadores(xlsx_path: str, df: pd.DataFrame):
    """
    df columns: Nome, Tipo, Retém, Tipo receita, Obs
    """
    wb = _open_wb(xlsx_path)
    ws = wb["Tomadores"]
    _clear_range(ws, TOMADORES_DATA_START, TOMADORES_DATA_END, TOMADORES_INPUT_COLS)

    col_map = {"Nome": 1, "Tipo": 2, "Retém": 3, "Tipo receita": 4, "Obs": 5}
    for i, (_, row) in enumerate(df.iterrows()):
        if i >= (TOMADORES_DATA_END - TOMADORES_DATA_START + 1):
            break
        r = TOMADORES_DATA_START + i
        for col_name, col_num in col_map.items():
            val = row.get(col_name, None)
            ws.cell(row=r, column=col_num).value = None if (pd.isna(val) if val is not None else False) else val

    _save_wb(wb, xlsx_path)


# ── Faturamento PJ ────────────────────────────────────────────────────────────
# Data rows: 7-14 (8 months), Client cols: 2-7 (6 clients max)

FAT_DATA_START = 7
FAT_DATA_END = 14
FAT_CLIENT_COLS = [2, 3, 4, 5, 6, 7]


def save_faturamento(xlsx_path: str, df: pd.DataFrame):
    """
    df: rows = months (Mai-Dez), columns = client names + "Mês"
    Writes only the billing values per client per month.
    """
    wb = _open_wb(xlsx_path)
    ws = wb["Faturamento PJ"]

    # Read current client names from row 6
    client_names = []
    for c in FAT_CLIENT_COLS:
        v = ws.cell(row=6, column=c).value
        client_names.append(str(v).strip() if v else "")

    # Write billing values
    for i, (_, row) in enumerate(df.iterrows()):
        if i >= (FAT_DATA_END - FAT_DATA_START + 1):
            break
        r = FAT_DATA_START + i
        for j, cl_name in enumerate(client_names):
            if cl_name and cl_name in df.columns:
                val = row.get(cl_name, None)
                col_num = FAT_CLIENT_COLS[j]
                ws.cell(row=r, column=col_num).value = None if (pd.isna(val) if val is not None else False) else val

    _save_wb(wb, xlsx_path)


# ── Monthly Pontuais ──────────────────────────────────────────────────────────
# Section 7️⃣ in monthly sheets: rows 68-83, Input cols: 1(Desc), 2(Cat), 3(Valor), 6(Status)
# Col 4 = separator, col 5 = formula

PONTUAIS_DATA_START = 68
PONTUAIS_DATA_END = 83
PONTUAIS_INPUT_COLS = [1, 2, 3, 6]


def save_month_revenues(xlsx_path: str, month_key: str, df: pd.DataFrame):
    """
    Updates Realizado (col 5) and Status (col 6) for revenues in a month sheet.
    df must have columns: Descrição, Realizado, Status.
    Matches rows by scanning col 1 for description substring, then writes cols 5 and 6.
    """
    import difflib

    wb = _open_wb(xlsx_path)
    if month_key not in wb.sheetnames:
        return
    ws = wb[month_key]

    for _, row in df.iterrows():
        desc_target = str(row.get("Descrição", "")).strip().lower()
        if not desc_target:
            continue

        best_row = -1
        best_score = 0.0
        for r in range(1, 80):
            cell_val = ws.cell(row=r, column=1).value
            if not cell_val:
                continue
            score = difflib.SequenceMatcher(None, str(cell_val).lower(), desc_target).ratio()
            if score > best_score:
                best_score = score
                best_row = r

        if best_row == -1 or best_score < 0.5:
            continue

        realizado = row.get("Realizado", None)
        status = row.get("Status", None)

        if realizado is not None:
            val = None if (pd.isna(realizado) if realizado is not None else False) else realizado
            ws.cell(row=best_row, column=5).value = val
        if status is not None:
            ws.cell(row=best_row, column=6).value = None if (pd.isna(status) if status is not None else False) else status

    _save_wb(wb, xlsx_path)


def save_pontuais_mes(xlsx_path: str, month_key: str, df: pd.DataFrame):
    """
    df columns: Descrição, Categoria, Valor, Status
    Clears pontuais rows and rewrites.
    """
    wb = _open_wb(xlsx_path)
    if month_key not in wb.sheetnames:
        return
    ws = wb[month_key]

    # Find actual section 7️⃣ row (may vary slightly between sheets)
    sec_row = PONTUAIS_DATA_START - 3  # default estimate
    for r in range(60, 100):
        v = str(ws.cell(row=r, column=1).value or "")
        if "7️⃣" in v:
            sec_row = r + 2  # header + 1 = data start
            break

    data_end = sec_row + 15  # max 16 pontuais

    _clear_range(ws, sec_row, data_end, PONTUAIS_INPUT_COLS)

    col_map = {"Descrição": 1, "Categoria": 2, "Valor": 3, "Status": 6}
    for i, (_, row) in enumerate(df.iterrows()):
        if i >= 16:
            break
        r = sec_row + i
        for col_name, col_num in col_map.items():
            val = row.get(col_name, None)
            ws.cell(row=r, column=col_num).value = None if (pd.isna(val) if val is not None else False) else val

    _save_wb(wb, xlsx_path)
