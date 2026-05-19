from dataclasses import dataclass, field
from typing import Optional
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from config import MONTHS, MONTH_LABELS


@dataclass
class MonthData:
    name: str
    label: str
    revenues: pd.DataFrame = field(default_factory=pd.DataFrame)
    pj_parcelas: pd.DataFrame = field(default_factory=pd.DataFrame)
    darf: pd.DataFrame = field(default_factory=pd.DataFrame)
    fixos: pd.DataFrame = field(default_factory=pd.DataFrame)
    parcelas_pessoais: pd.DataFrame = field(default_factory=pd.DataFrame)
    pontuais: pd.DataFrame = field(default_factory=pd.DataFrame)
    summary: dict = field(default_factory=dict)


@dataclass
class PainelData:
    kpis: dict
    monthly_summary: pd.DataFrame
    expense_composition: pd.DataFrame
    faturamento: pd.DataFrame
    darf_by_month: pd.DataFrame
    fixos: pd.DataFrame
    pj_parcelas_config: pd.DataFrame
    parcelas_pessoais_config: pd.DataFrame
    months: dict
    config: dict


def _cell_val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return v


def _str(v):
    if v is None:
        return ""
    return str(v).strip()


def _num(v, default=0.0):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _parse_overview(ws):
    kpis = {}
    # KPI row is at row 5 (0-indexed row 4): saldo, fixo_mensal, reserva_alvo, runway
    # Row 6 has break_even and saldo_dez
    try:
        kpis["saldo_inicial"] = _num(_cell_val(ws, 5, 2))
        kpis["fixo_mensal"] = _num(_cell_val(ws, 5, 5))
        kpis["reserva_alvo"] = _num(_cell_val(ws, 5, 8))
        kpis["runway"] = _num(_cell_val(ws, 5, 11))
        kpis["break_even"] = _num(_cell_val(ws, 7, 2))
        kpis["progresso_reserva"] = _num(_cell_val(ws, 7, 4))
        kpis["saldo_dez"] = _num(_cell_val(ws, 7, 11))
    except Exception:
        pass

    # Monthly summary table — rows 10-18 (Excel rows 10..18), header at row 9
    months_rows = []
    col_names = ["Mês", "Receitas", "Parc. PJ", "DARF", "Fixos", "Parcelas", "Pontuais",
                 "Total Despesas", "Saldo do Mês", "Saldo Acumulado", "% Comprometido", "Status"]
    for r in range(10, 20):
        row_vals = [_cell_val(ws, r, c) for c in range(1, 13)]
        if row_vals[0] and str(row_vals[0]).strip():
            months_rows.append(row_vals)

    monthly_df = pd.DataFrame(months_rows, columns=col_names) if months_rows else pd.DataFrame()

    # Expense composition — rows 22-27
    exp_rows = []
    for r in range(22, 28):
        cat = _str(_cell_val(ws, r, 1))
        total = _num(_cell_val(ws, r, 2))
        pct = _num(_cell_val(ws, r, 3))
        if cat and cat not in ("Categoria", "TOTAL DESPESAS"):
            exp_rows.append({"Categoria": cat, "Total": total, "Pct": pct})
    exp_df = pd.DataFrame(exp_rows) if exp_rows else pd.DataFrame()

    return kpis, monthly_df, exp_df


def _parse_config(ws):
    cfg = {}
    for r in range(1, 26):
        k = _str(_cell_val(ws, r, 1))
        v = _cell_val(ws, r, 2)
        if k:
            cfg[k] = v
    return cfg


def _parse_faturamento(ws):
    # Billing section: rows 6-14, clients in columns 2-6+ (row 5 = header, row 6 = first data)
    # Client names in row 6 cols 2..7
    clients = []
    for c in range(2, 8):
        v = _str(_cell_val(ws, 6, c))
        if v and v not in ("NaN", ""):
            clients.append(v)

    rows = []
    for r in range(7, 15):
        mes = _str(_cell_val(ws, r, 1))
        if not mes or mes.lower() in ("nan", ""):
            continue
        row = {"Mês": mes}
        for i, cl in enumerate(clients):
            row[cl] = _num(_cell_val(ws, r, i + 2))
        row["Total"] = _num(_cell_val(ws, r, 12))
        rows.append(row)
    fat_df = pd.DataFrame(rows) if rows else pd.DataFrame()

    # DARF table: starts around row 17
    darf_rows = []
    darf_cols = ["Competência", "Mês Pagamento", "PIS", "COFINS", "IRPJ", "CSLL", "ISS",
                 "Total Bruto", "(-) Retenções", "DARF a Pagar"]
    for r in range(18, 27):
        comp = _str(_cell_val(ws, r, 1))
        if not comp or comp.lower() in ("nan", ""):
            continue
        row_vals = [_str(_cell_val(ws, r, c)) for c in range(1, 11)]
        darf_rows.append(row_vals)
    darf_df = pd.DataFrame(darf_rows, columns=darf_cols) if darf_rows else pd.DataFrame()

    return fat_df, darf_df


def _parse_fixos(ws):
    rows = []
    for r in range(5, 30):
        desc = _str(_cell_val(ws, r, 1))
        if not desc or desc.lower() in ("nan", "total mensal fixo"):
            continue
        cat = _str(_cell_val(ws, r, 2))
        val = _num(_cell_val(ws, r, 3))
        status = _str(_cell_val(ws, r, 4))
        obs = _str(_cell_val(ws, r, 5))
        if val > 0:
            rows.append({"Descrição": desc, "Categoria": cat, "Valor": val,
                         "Status": status, "Obs": obs})
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _parse_pj_parcelas(ws):
    rows = []
    for r in range(5, 20):
        desc = _str(_cell_val(ws, r, 1))
        if not desc or desc.lower() in ("nan", "total compromissado"):
            continue
        val = _num(_cell_val(ws, r, 2))
        n_parc = _num(_cell_val(ws, r, 3))
        mes_ini = _num(_cell_val(ws, r, 4))
        ano_ini = _num(_cell_val(ws, r, 5))
        mes_fim = _str(_cell_val(ws, r, 6))
        total = _num(_cell_val(ws, r, 7))
        status = _str(_cell_val(ws, r, 8))
        obs = _str(_cell_val(ws, r, 9))
        if val > 0:
            rows.append({"Descrição": desc, "Parcela": val, "N Parcelas": int(n_parc),
                         "Mês Início": int(mes_ini), "Ano Início": int(ano_ini),
                         "Mês Final": mes_fim, "Total": total, "Status": status, "Obs": obs})
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _parse_parcelas_pessoais(ws):
    rows = []
    for r in range(5, 25):
        desc = _str(_cell_val(ws, r, 1))
        if not desc or desc.lower() in ("nan", "total em amortização"):
            continue
        cat = _str(_cell_val(ws, r, 2))
        val = _num(_cell_val(ws, r, 3))
        n_parc = _num(_cell_val(ws, r, 4))
        mes_ini = _num(_cell_val(ws, r, 5))
        ano_ini = _num(_cell_val(ws, r, 6))
        mes_fim = _str(_cell_val(ws, r, 7))
        total = _num(_cell_val(ws, r, 8))
        obs = _str(_cell_val(ws, r, 9))
        if val > 0:
            rows.append({"Descrição": desc, "Categoria": cat, "Parcela": val,
                         "N Parcelas": int(n_parc), "Mês Início": int(mes_ini),
                         "Ano Início": int(ano_ini), "Mês Final": mes_fim,
                         "Total": total, "Obs": obs})
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _find_section_rows(ws, max_row=150):
    """Return dict mapping section number (1-8) to the row where that section starts."""
    sections = {}
    markers = {
        "1️⃣": 1, "2️⃣": 2, "3️⃣": 3, "4️⃣": 4,
        "5️⃣": 5, "6️⃣": 6, "7️⃣": 7, "8️⃣": 8,
    }
    for r in range(1, max_row):
        v = _str(_cell_val(ws, r, 1))
        for emoji, num in markers.items():
            if emoji in v:
                sections[num] = r
                break
    return sections


def _parse_section_table(ws, header_row, end_row, ncols=6):
    """Parse a section's table: header_row+1 = col names, header_row+2..end_row = data."""
    if header_row is None or end_row is None or end_row <= header_row + 1:
        return pd.DataFrame()
    cols = [_str(_cell_val(ws, header_row + 1, c)) or f"col{c}" for c in range(1, ncols + 1)]
    rows = []
    for r in range(header_row + 2, end_row):
        row_vals = [_cell_val(ws, r, c) for c in range(1, ncols + 1)]
        first = _str(row_vals[0])
        if not first or first.lower() == "nan":
            continue
        # Stop at totals/summary rows
        if any(x in first for x in ("📥", "💰", "TOTAL", "SALDO", "subtotal")):
            break
        rows.append(row_vals)
    return pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame()


def _get_summary_value(ws, section_start, label_fragment, value_col=2, search_rows=20):
    for r in range(section_start, section_start + search_rows):
        v = _str(_cell_val(ws, r, 1))
        if label_fragment.lower() in v.lower():
            return _num(_cell_val(ws, r, value_col))
    return 0.0


def _parse_month_sheet(ws, name, label):
    sections = _find_section_rows(ws)

    def end_of(sec_num):
        next_secs = sorted(v for k, v in sections.items() if k > sec_num)
        return next_secs[0] if next_secs else ws.max_row

    # 1. Receitas
    revenues = pd.DataFrame()
    if 1 in sections:
        s = sections[1]
        e = end_of(1)
        cols = ["Descrição", "Tipo", "Data", "Previsto", "Realizado", "Status"]
        rows = []
        for r in range(s + 2, e):
            desc = _str(_cell_val(ws, r, 1))
            if not desc or "TOTAL" in desc or "CAIXA" in desc:
                continue
            rows.append({
                "Descrição": desc,
                "Tipo": _str(_cell_val(ws, r, 2)),
                "Data": _str(_cell_val(ws, r, 3)),
                "Previsto": _num(_cell_val(ws, r, 4)),
                "Realizado": _num(_cell_val(ws, r, 5)),
                "Status": _str(_cell_val(ws, r, 6)),
            })
        revenues = pd.DataFrame(rows) if rows else pd.DataFrame()

    # 3. PJ Parcelamentos
    pj_parcelas = pd.DataFrame()
    if 3 in sections:
        s = sections[3]
        e = end_of(3)
        rows = []
        for r in range(s + 2, e):
            desc = _str(_cell_val(ws, r, 1))
            if not desc:
                continue
            val = _num(_cell_val(ws, r, 5))
            rows.append({
                "Descrição": desc,
                "Parcela Atual": _str(_cell_val(ws, r, 2)),
                "Valor": val,
                "Status": _str(_cell_val(ws, r, 6)),
            })
        pj_parcelas = pd.DataFrame(rows) if rows else pd.DataFrame()

    # 4. DARF
    darf = pd.DataFrame()
    if 4 in sections:
        s = sections[4]
        e = end_of(4)
        rows = []
        for r in range(s + 2, e):
            desc = _str(_cell_val(ws, r, 1))
            if not desc or "TOTAL" in desc.upper():
                continue
            val = _num(_cell_val(ws, r, 4))
            rows.append({
                "Descrição": desc,
                "Competência": _str(_cell_val(ws, r, 2)),
                "Vencimento": _str(_cell_val(ws, r, 3)),
                "Valor": val,
                "Status": _str(_cell_val(ws, r, 5)),
            })
        darf = pd.DataFrame(rows) if rows else pd.DataFrame()

    # 5. Fixos Pessoais
    fixos = pd.DataFrame()
    if 5 in sections:
        s = sections[5]
        e = end_of(5)
        rows = []
        for r in range(s + 2, e):
            desc = _str(_cell_val(ws, r, 1))
            if not desc or "TOTAL" in desc.upper():
                continue
            val = _num(_cell_val(ws, r, 3))
            if val > 0:
                rows.append({
                    "Descrição": desc,
                    "Categoria": _str(_cell_val(ws, r, 2)),
                    "Valor": val,
                })
        fixos = pd.DataFrame(rows) if rows else pd.DataFrame()

    # 6. Parcelas Pessoais
    parcelas_pessoais = pd.DataFrame()
    if 6 in sections:
        s = sections[6]
        e = end_of(6)
        rows = []
        for r in range(s + 2, e):
            desc = _str(_cell_val(ws, r, 1))
            if not desc or "TOTAL" in desc.upper():
                continue
            val = _num(_cell_val(ws, r, 3))
            if val > 0:
                rows.append({
                    "Descrição": desc,
                    "Parcela Atual": _str(_cell_val(ws, r, 2)),
                    "Valor": val,
                    "Status": _str(_cell_val(ws, r, 4)),
                })
        parcelas_pessoais = pd.DataFrame(rows) if rows else pd.DataFrame()

    # 7. Gastos Pontuais
    pontuais = pd.DataFrame()
    if 7 in sections:
        s = sections[7]
        e = end_of(7)
        rows = []
        for r in range(s + 2, e):
            desc = _str(_cell_val(ws, r, 1))
            if not desc or "TOTAL" in desc.upper():
                continue
            val = _num(_cell_val(ws, r, 3))
            rows.append({
                "Descrição": desc,
                "Categoria": _str(_cell_val(ws, r, 2)),
                "Valor": val,
                "Status": _str(_cell_val(ws, r, 4)),
            })
        pontuais = pd.DataFrame(rows) if rows else pd.DataFrame()

    # 8. Resumo — find summary values
    summary = {}
    if 8 in sections:
        s = sections[8]
        for r in range(s, min(s + 30, ws.max_row)):
            label = _str(_cell_val(ws, r, 1)).lower()
            val2 = _cell_val(ws, r, 2)
            val3 = _cell_val(ws, r, 3)
            if "receita" in label and "total" in label:
                summary["receita_total"] = _num(val2)
            elif "total despesa" in label or "total gasto" in label:
                summary["total_despesas"] = _num(val2)
            elif "saldo do mês" in label or "saldo final" in label:
                summary["saldo_mes"] = _num(val2)
            elif "saldo acumulado" in label:
                summary["saldo_acumulado"] = _num(val2)
            elif "comprometido" in label or "% comprometido" in label:
                summary["pct_comprometido"] = _num(val2)

    return MonthData(
        name=name, label=label,
        revenues=revenues, pj_parcelas=pj_parcelas, darf=darf,
        fixos=fixos, parcelas_pessoais=parcelas_pessoais,
        pontuais=pontuais, summary=summary,
    )


@st.cache_data(ttl=60)
def load(xlsx_path: str) -> PainelData:
    wb = load_workbook(xlsx_path, data_only=True)

    kpis, monthly_summary, expense_comp = _parse_overview(wb["Visão Geral"])
    cfg = _parse_config(wb["Configurações"])
    fat_df, darf_df = _parse_faturamento(wb["Faturamento PJ"])
    fixos_df = _parse_fixos(wb["Fixos Pessoais"])
    pj_parc_df = _parse_pj_parcelas(wb["PJ Parcelamentos"])
    parc_pes_df = _parse_parcelas_pessoais(wb["Parcelas Pessoais"])

    months = {}
    for name, label in zip(MONTHS, MONTH_LABELS):
        if name in wb.sheetnames:
            months[name] = _parse_month_sheet(wb[name], name, label)

    return PainelData(
        kpis=kpis,
        monthly_summary=monthly_summary,
        expense_composition=expense_comp,
        faturamento=fat_df,
        darf_by_month=darf_df,
        fixos=fixos_df,
        pj_parcelas_config=pj_parc_df,
        parcelas_pessoais_config=parc_pes_df,
        months=months,
        config=cfg,
    )
