# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import OrderedDict, Counter

# ----- Protocolos EXISTENTES (do inventario layout novo): (Protocolo, Especialidade) -----
existentes = []
def add(esp, lista):
    for p in lista:
        existentes.append((p, esp))

add("Gastroenterologia", ["HDA — Hemorragia Digestiva Alta","DUP — Doença Ulcerosa Péptica","Diarreia Aguda","Diarreia Crônica","SII — Síndrome do Intestino Irritável","Elevação de Enzimas Hepáticas","Nutrição Enteral em Paciente Crítico","Constipação Intestinal","Dispepsia","Cirrose Hepática no PS","Ingestão Cáustica","Doença de Crohn","DRGE","Hepatite Alcoólica","Retocolite Ulcerativa","Síndrome de Boerhaave"])
add("Cardiologia", ["Taquiarritmias no PS","SCC — Síndrome Coronariana Crônica","Dissecção Aórtica Aguda","HAS no Pronto-Socorro","EAP — Edema Agudo de Pulmão","Cardiomiopatias","Bradiarritmias no PS","IC Aguda — Insuficiência Cardíaca","Tamponamento Cardíaco","Síndrome Coronariana Aguda","Síncope no PS","Parada Cardiorrespiratória / ACLS","ECG — Oclusão e Risco","Endocardite Infecciosa","Miocardite e Pericardite"])
add("Clínica Médica", ["Acidente Ofídico","Investigação de Plaquetopenia","Urticária Aguda","Espasmo Hemifacial e Paraparesia Espástica","Intoxicações Exógenas","Sepse e Choque Séptico","Anafilaxia"])
add("Infectologia", ["Zika","Esquistossomose","Furúnculo / HS","Hepatites Virais","Febre Tifoide","Herpes Zoster","Dengue no PS","Exposição Ocupacional — PEP HIV-HBV-HCV","HIV e Infecções Oportunistas","Antibioticoterapia Empírica","Meningite e Encefalite no Adulto","Chikungunya","Dengue","Febre Maculosa","Coqueluche","Caxumba","Leptospirose","Malária","Herpes Simples","Influenza / Síndrome Gripal","Tuberculose","Celulite e Erisipela","Ectoparasitoses","Furúnculo, Carbúnculo e Hidradenite","Sífilis","Úlceras Genitais"])
add("Pneumologia", ["Hemoptise","DPOC Exacerbado","Crise Asmática","SDRA","Pneumonia (PAC) no Adulto","Pneumotórax Espontâneo","Insuficiência Respiratória Aguda / VNI","Tabagismo","TEP — Tromboembolismo Pulmonar"])
add("Endocrinologia", ["SAHOS","Hiperglicemia Hospitalar","Hipoglicemia","Coma Mixedematoso e Feocromocitoma","Crise Tireotóxica","Cetoacidose Diabética","Insuficiência Adrenal","Síndrome de Cushing","Tireoidites"])
add("Psiquiatria", ["Risco de Suicídio — Avaliação","Ansiedade e Crise de Pânico","Síndrome de Abstinência Alcoólica","Síndrome Serotoninérgica","Delirium no Idoso","Agitação Psicomotora"])
add("Nefrologia", ["LRA — Lesão Renal Aguda e NTA","SHU — Síndrome Hemolítico-Urêmica","Hipernatremia","Hiponatremia","Gasometria — Interpretação Prática","Hipercalemia","Diálise de Urgência","Distúrbios de Magnésio e Fósforo","Distúrbios Ácido-Base","Distúrbios do Cálcio","Doença Renal Crônica","GNRP","Hipocalemia","Nefrite Intersticial Aguda","Rabdomiólise","Síndrome Hepatorrenal","Síndrome Nefrítica","Síndrome Nefrótica"])
add("Neurologia", ["Infarto Medular e Diplopia Binocular","Tremor Essencial e Neuralgia Pós-Herpética","Estado de Mal Epiléptico e AIT","Crise Miastênica e Síndrome Extrapiramidal","AVC Isquêmico e Hemorrágico","Neuralgias e HII","Tremor no PS e Fraqueza Muscular Aguda","Neuropatia Diabética e Túnel do Carpo","Cefaleia Tensional, MOH e Arterite Temporal","Doença de Parkinson e RCVS","PRES — Encefalopatia Posterior Reversível","Trombose Venosa Cerebral","Síndrome da Cauda Equina","Síndrome Neuroléptica Maligna","Cefaleia Tensional (v1)","Cefaleia em Salvas","Enxaqueca","AGT — Amnésia Global Transitória","Cefaleia no Adulto","Síndrome de Guillain-Barré","Paralisia Facial Periférica","Coma e Rebaixamento de Consciência","TCE no Adulto","AVC (apostila AVC)"])
add("Procedimentos", ["Procedimentos — Toracocentese, Drenagem, Paracentese, IO, Traqueostomia","Retenção Urinária e Sondagem","Sutura e Manejo de Feridas","Bloqueios Anestésicos no PS","Sequência Rápida de Intubação","Acesso Venoso Central","Paciente Traqueostomizado"])
add("Cirurgia Geral", ["Abdome Agudo — Obstrutivo, Inflamatório, Perfurativo e Megacólon","Abdome Vascular — Isquemia Mesentérica e Anestesia Local","Apendicite","Torção Testicular","Cólica Renal e Litíase","Colecistite Aguda","Coledocolitíase","Colangite Aguda","Pancreatite Aguda","Pancreatite Crônica","Colite Isquêmica","Câncer Colorretal","Acalasia","Corpo Estranho Digestivo","Doença Hemorroidária","Diverticulite Aguda","Fissura Anal","Fístula e Abscesso Perianal","Prolapso Retal","Complicações de Ferida Operatória","Esôfago de Barrett","Gastroparesia","Hérnia de Hiato","Lesão Esofágica Cáustica","Perfuração Esofágica","Úlcera Péptica","Febre Pós-Operatória","Hérnias da Parede Abdominal","Íleo Paralítico","Drenagem de Abscesso Cutâneo","Fasciíte Necrotizante","Lesão por Pressão","Mordeduras e Ferimentos","Queimaduras","Complicações de Traqueostomia","Cricotireoidostomia","Derrame Pleural","Drenagem de Tórax","Mediastinite Aguda","Paracentese","Pós-operatório de Cirurgia Torácica"])
add("Trauma", ["Trauma em Populações Especiais — Idoso, Criança, Gestante","Trauma de Via Aérea e Choque Elétrico","Trauma Torácico","Fratura de Costelas e Contusão Pulmonar","TCE Grave e HIC","Choque Hemorrágico e Transfusão","Afogamento e Hipotermia","Atendimento Inicial ao Trauma","FAST no Trauma","Ferimento Descolante","Hemotórax","Pneumotórax","Trauma Abdominal","Trauma Cervical","Trauma Gastrointestinal","Trauma de Face","Trauma Vascular de Extremidades","Trauma de Pelve"])
add("Vascular", ["TVP — Trombose Venosa Profunda","Aneurisma de Aorta Abdominal","Insuficiência Venosa Crônica","Oclusão Arterial Aguda","Oclusão Arterial Crônica","Pé Diabético","Pseudoaneurisma","Síndrome da Veia Cava Superior"])
add("Ortopedia", ["Fratura Exposta","Fasciíte Plantar","Cervicalgia","Lombalgia","Costocondrite","Entorse de Tornozelo","Trauma Raquimedular","Luxação de Ombro","Entorse de Joelho","Bursite","Síndrome Compartimental","Artrite Séptica","Crise de Gota / Artrite","Osteomielite"])
add("Pediatria", ["Pneumonia (PAC) em Pediatria","Bronquiolite","TSS Estafilocócica","Doença Meningocócica (Pediatria)","Doenças Exantemáticas (Pediatria)","ITU (Pediatria)","Mão-Pé-Boca","Dengue (Pediatria)","Impetigo","Febre Amarela (Pediatria)","DAG — Diarreia Aguda (Pediatria)","Sepse Neonatal","Maus-Tratos Infantil","Laringite e Estridor","CAD na Criança","Convulsão Febril","Acidente Peçonhento (Pediatria)","Asma (Pediatria)","Crupe e Epiglotite","Desidratação (Pediatria)","Diarreia na Criança","Dor Abdominal Pediátrica","PCR na Infância","TCE na Infância","Trauma Ortopédico (Pediatria)","VMI em Pediatria"])
add("Hematologia", ["Anticoagulação e Reversão","PTI — Púrpura Trombocitopênica","Crise Falciforme","CIVD","Anemia","Pancitopenia","Reação Transfusional Febril"])
add("Oftalmologia", ["Oclusão de Artéria/Veia da Retina","Endoftalmite","Descolamento de Retina","Conjuntivite","Corpo Estranho Ocular","Glaucoma Agudo","Trauma Ocular"])
add("Cirurgia Cabeça e Pescoço", ["Abscesso Cervical","Angina de Ludwig","Complicações de Tireoidectomia"])
add("Otorrinolaringologia", ["Vertigem e VPPB","Surdez Súbita","Corpo Estranho Auditivo","Celulite Facial","Epistaxe","Faringoamigdalite Aguda","Otite Média Aguda","Sinusite Aguda"])
add("Ginecologia e Obstetrícia", ["Trabalho de Parto Prematuro","Emergências Hipertensivas na Gestação","Torção de Ovário","Aborto Infectado","Asma na Gestante","Diabetes Gestacional","Doença Inflamatória Pélvica","Hemorragia Pós-Parto","ITU na Gestante","Mastite","Parto Normal","Pré-eclâmpsia e Eclâmpsia","Sangramento de Primeiro Trimestre","Sangramento Uterino Anormal","Sepse Puerperal","Síndrome HELLP","Vaginose Bacteriana"])
add("Oncologia", ["Mucosite e Toxicidades da Quimio","Compressão Medular Neoplásica","Carcinoma Hepatocelular","Dor Oncológica e Opioides","HIC na Malignidade","Hipercalcemia da Malignidade","Neoplasia de Sítio Primário Desconhecido","Neoplasia de Testículo","Neutropenia Febril","Rastreamento de Câncer","Síndrome de Lise Tumoral","TEV no Câncer"])
add("Odontologia", ["Pericoronarite","Hemorragia Pós-Exodontia","Avulsão Dentária","Abscesso Dentoalveolar","Gengivoestomatite Herpética"])

# ----- 38 NOVOS protocolos (produzidos): (Protocolo, Especialidade) -----
novos = [
    ("Crise Hipertensiva — Urgência vs Emergência","Cardiologia"),
    ("Fibrilação Atrial no PS — Ritmo vs Frequência","Cardiologia"),
    ("Valvopatias Agudas no PS","Cardiologia"),
    ("Crise de Hipertensão Pulmonar","Pneumologia"),
    ("Aspiração e Pneumonite Química","Pneumologia"),
    ("Distúrbios da Glicemia no Etilista","Clínica Médica"),
    ("Síndrome de Realimentação","Clínica Médica"),
    ("Dor Aguda no PS — Escada Analgésica","Clínica Médica"),
    ("Delirium vs Demência vs Depressão no Idoso","Clínica Médica"),
    ("Anticoagulação Periprocedimento","Clínica Médica"),
    ("Estado Hiperglicêmico Hiperosmolar (EHH)","Endocrinologia"),
    ("Hipo e Hipercalcemia Severas","Endocrinologia"),
    ("Encefalopatia Hepática","Gastroenterologia"),
    ("Pancreatite Biliar — Manejo Integrado","Gastroenterologia"),
    ("Nefropatia Induzida por Contraste","Nefrologia"),
    ("Mielite e Compressão Medular Não-Traumática","Neurologia"),
    ("Sepse de Foco Urinário e Pielonefrite","Infectologia"),
    ("Profilaxia Antirrábica Pós-Exposição","Infectologia"),
    ("Tétano — Profilaxia e Manejo","Infectologia"),
    ("Obstrução Intestinal — Bridas vs Tumor","Cirurgia Geral"),
    ("Trauma Raquimedular e Choque Neurogênico","Trauma"),
    ("Manejo do Grande Queimado","Trauma"),
    ("Fratura de Quadril no Idoso","Ortopedia"),
    ("Síndrome do Túnel do Carpo Agudo","Ortopedia"),
    ("Apneia e BRUE no Lactente","Pediatria"),
    ("Intoxicação Acidental na Criança","Pediatria"),
    ("Neutropenia Febril Não-Oncológica","Hematologia"),
    ("Transfusão Maciça — Protocolo","Hematologia"),
    ("Cetoacidose Diabética na Gestação","Ginecologia e Obstetrícia"),
    ("Gravidez Ectópica Rota","Ginecologia e Obstetrícia"),
    ("Celulite Orbitária vs Pré-Septal","Oftalmologia"),
    ("Obstrução de Via Aérea Superior no Adulto","Otorrinolaringologia"),
    ("Catatonia — Reconhecimento e Manejo","Psiquiatria"),
    ("Intoxicação por Benzodiazepínicos e Opioides","Psiquiatria"),
    ("Celulite Odontogênica com Repercussão Sistêmica","Odontologia"),
    ("Trombose de Acesso para Hemodiálise","Vascular"),
    ("Emergências de Imunoterapia (irAEs)","Oncologia"),
    ("Sedação e Analgesia para Procedimentos no PS","Procedimentos"),
]

# ----- Sugestoes (agora PRODUZIDAS): (Especialidade, Protocolo Sugerido, Justificativa) -----
sugestoes = [
    ("Cardiologia","Crise Hipertensiva vs Urgência/Emergência Hipertensiva","Diferenciação prática frequente no PS, ainda não isolada"),
    ("Cardiologia","Fibrilação Atrial — Controle de Ritmo vs Frequência","Subtópico de alta prevalência dentro de taquiarritmias"),
    ("Cardiologia","Valvopatias Agudas no PS","Lacuna entre cardiomiopatias e endocardite"),
    ("Pneumologia","Crise de Hipertensão Pulmonar","Emergência respiratória não coberta"),
    ("Pneumologia","Aspiração e Pneumonite Química","Comum em idosos e rebaixados"),
    ("Clínica Médica","Distúrbios da Glicemia no Etilista","Cruzamento endócrino-toxicológico frequente"),
    ("Clínica Médica","Síndrome de Realimentação","Relevante em desnutridos e etilistas internados"),
    ("Clínica Médica","Manejo da Dor Aguda no PS — Escada Analgésica","Protocolo transversal de alto uso"),
    ("Endocrinologia","Estado Hiperglicêmico Hiperosmolar (EHH)","Par natural da CAD, ausente na pasta"),
    ("Endocrinologia","Hipo e Hipercalcemia Severas (manejo agudo)","Complementa distúrbios do cálcio da Nefro"),
    ("Nefrologia","Manejo de Contraste e Nefropatia Induzida","Decisão frequente em PS com TC"),
    ("Neurologia","Mielite e Compressão Medular Não-Traumática","Emergência neurológica ausente"),
    ("Infectologia","Sepse de Foco Urinário / Pielonefrite","Alta prevalência, foco específico"),
    ("Infectologia","Profilaxia Antirrábica Pós-Exposição","Protocolo de notificação frequente"),
    ("Infectologia","Tétano — Profilaxia e Manejo","Decisão comum em ferimentos no PS"),
    ("Gastroenterologia","Encefalopatia Hepática","Complemento natural de cirrose no PS"),
    ("Gastroenterologia","Pancreatite Biliar — Manejo Integrado","Cruzamento Gastro/Cirurgia"),
    ("Cirurgia Geral","Obstrução Intestinal por Bridas vs Tumor","Subtópico de alta decisão clínica"),
    ("Trauma","Trauma Raquimedular — Choque Neurogênico","Complementa trauma cervical/RM"),
    ("Trauma","Manejo de Grande Queimado (ATLS/transferência)","Critérios de transferência ausentes"),
    ("Ortopedia","Fratura de Quadril no Idoso","Altíssima prevalência, ausente"),
    ("Ortopedia","Síndrome do Túnel do Carpo Agudo","Complemento ambulatorial"),
    ("Pediatria","Apneia e ALTE/BRUE no Lactente","Emergência pediátrica ausente"),
    ("Pediatria","Intoxicação Acidental na Criança","Frequente em PS pediátrico"),
    ("Hematologia","Neutropenia Febril Não-Oncológica","Complementa item da Oncologia"),
    ("Hematologia","Transfusão Maciça — Protocolo","Cruzamento com trauma"),
    ("Ginecologia e Obstetrícia","Cetoacidose Diabética na Gestação","Emergência metabólica obstétrica"),
    ("Ginecologia e Obstetrícia","Gravidez Ectópica Rota","Emergência cirúrgica ausente"),
    ("Oftalmologia","Celulite Orbitária vs Pré-Septal","Diferencial crítico no PS"),
    ("Otorrinolaringologia","Obstrução de Via Aérea Superior no Adulto","Emergência ORL ausente"),
    ("Psiquiatria","Catatonia — Reconhecimento e Manejo","Diferencial de SNM/rebaixamento"),
    ("Psiquiatria","Intoxicação por Benzodiazepínicos e Opioides","Complemento toxicológico psiquiátrico"),
    ("Odontologia","Celulite Odontogênica com Repercussão Sistêmica","Cruzamento com cabeça e pescoço"),
    ("Vascular","Trombose de Acesso para Hemodiálise","Lacuna em pacientes renais crônicos"),
    ("Oncologia","Emergências de Imunoterapia (irAEs)","Tema crescente e ausente"),
    ("Geral / Transversal","Sedação e Analgesia para Procedimentos no PS","Protocolo transversal de alto uso"),
    ("Geral / Transversal","Manejo de Anticoagulação Periprocedimento","Decisão recorrente em várias especialidades"),
    ("Geral / Transversal","Delirium vs Demência vs Depressão no Idoso","Diferencial transversal"),
]

# ----- ordem das especialidades (como no inventario) -----
ordem = ["Gastroenterologia","Cardiologia","Clínica Médica","Infectologia","Pneumologia","Endocrinologia","Psiquiatria","Nefrologia","Neurologia","Procedimentos","Cirurgia Geral","Trauma","Vascular","Ortopedia","Pediatria","Hematologia","Oftalmologia","Cirurgia Cabeça e Pescoço","Otorrinolaringologia","Ginecologia e Obstetrícia","Oncologia","Odontologia"]

todos = existentes + novos
cont = Counter(esp for _,esp in todos)

# ===== Construir o workbook =====
wb = openpyxl.Workbook()
NAVY="1F3864"; TEAL="2E8B8B"; WHITE="FFFFFF"; LIGHT="EAF1F1"; GREEN_F="C6EFCE"
hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=NAVY)
cell_font = Font(name="Arial", size=10)
new_fill = PatternFill("solid", fgColor=GREEN_F)
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
    ws.freeze_panes = "A2"

# --- Aba 1: Contagem por Especialidade ---
ws1 = wb.active; ws1.title = "Contagem por Especialidade"
ws1.append(["Especialidade","Nº de Protocolos"])
for esp in ordem:
    ws1.append([esp, cont.get(esp,0)])
ws1.append(["TOTAL", sum(cont.values())])
style_header(ws1,2)
for r in range(2, ws1.max_row+1):
    ws1.cell(r,1).font = cell_font; ws1.cell(r,1).alignment = left; ws1.cell(r,1).border=border
    ws1.cell(r,2).font = cell_font; ws1.cell(r,2).alignment = center; ws1.cell(r,2).border=border
tot_r = ws1.max_row
ws1.cell(tot_r,1).font = Font(name="Arial", bold=True, size=10)
ws1.cell(tot_r,2).font = Font(name="Arial", bold=True, size=10)
ws1.cell(tot_r,1).fill = PatternFill("solid", fgColor=LIGHT)
ws1.cell(tot_r,2).fill = PatternFill("solid", fgColor=LIGHT)
ws1.column_dimensions["A"].width=32; ws1.column_dimensions["B"].width=18

# --- Aba 2: Protocolos ---
ws2 = wb.create_sheet("Protocolos")
ws2.append(["Protocolo","Especialidade","Status"])
novos_set = set(p for p,_ in novos)
for esp in ordem:
    for p, e in todos:
        if e == esp:
            status = "NOVO (jun/2026)" if p in novos_set else "Existente"
            ws2.append([p, e, status])
style_header(ws2,3)
for r in range(2, ws2.max_row+1):
    for c in range(1,4):
        cell = ws2.cell(r,c); cell.font=cell_font; cell.border=border
        cell.alignment = left if c==1 else center
    if ws2.cell(r,3).value and ws2.cell(r,3).value.startswith("NOVO"):
        for c in range(1,4): ws2.cell(r,c).fill = new_fill
ws2.column_dimensions["A"].width=58; ws2.column_dimensions["B"].width=26; ws2.column_dimensions["C"].width=18

# --- Aba 3: Sugestões (Produzidas) ---
ws3 = wb.create_sheet("Sugestões")
ws3.append(["Especialidade","Protocolo Sugerido","Justificativa","Status"])
for esp, prot, just in sugestoes:
    ws3.append([esp, prot, just, "PRODUZIDO"])
style_header(ws3,4)
for r in range(2, ws3.max_row+1):
    for c in range(1,5):
        cell = ws3.cell(r,c); cell.font=cell_font; cell.border=border
        cell.alignment = left if c in (2,3) else center
    ws3.cell(r,4).fill = new_fill
ws3.column_dimensions["A"].width=24; ws3.column_dimensions["B"].width=42; ws3.column_dimensions["C"].width=46; ws3.column_dimensions["D"].width=14

out = "SM24_Inventario_Protocolos_atualizado.xlsx"
wb.save(out)

print("Existentes:", len(existentes), "| Novos:", len(novos), "| Total:", len(todos))
print("Contagem por especialidade:")
for esp in ordem:
    print(f"  {esp}: {cont.get(esp,0)}")
print("Arquivo:", out)
import os
print("Tamanho xlsx:", os.path.getsize(out), "bytes")
